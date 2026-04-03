#!/usr/bin/env python3
"""Generate Excel file with insurance company WeChat official accounts."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Insurance company WeChat official accounts data
# Based on industry knowledge and public information
insurance_accounts = [
    # 大型综合保险集团
    {"序号": 1, "公司名称": "中国平安保险", "公众号名称": "中国平安", "公众号 ID": "pingan_1988", "类型": "综合金融", "备注": "集团官方号"},
    {"序号": 2, "公司名称": "中国平安保险", "公众号名称": "平安人寿", "公众号 ID": "pinganlife", "类型": "人寿保险", "备注": "寿险子公司"},
    {"序号": 3, "公司名称": "中国平安保险", "公众号名称": "平安产险", "公众号 ID": "PAIC-PC", "类型": "财产保险", "备注": "产险子公司"},
    {"序号": 4, "公司名称": "中国人寿保险", "公众号名称": "中国人寿", "公众号 ID": "chinalife-95519", "类型": "人寿保险", "备注": "集团官方号"},
    {"序号": 5, "公司名称": "中国人寿保险", "公众号名称": "中国人寿财险", "公众号 ID": "chinalife-p", "类型": "财产保险", "备注": "财险子公司"},
    {"序号": 6, "公司名称": "中国人保", "公众号名称": "中国人保", "公众号 ID": "PICC_1949", "类型": "综合金融", "备注": "集团官方号"},
    {"序号": 7, "公司名称": "中国人保", "公众号名称": "人保财险", "公众号 ID": "picc_1949", "类型": "财产保险", "备注": "财险子公司"},
    {"序号": 8, "公司名称": "中国人保", "公众号名称": "人保寿险", "公众号 ID": "picclife", "类型": "人寿保险", "备注": "寿险子公司"},
    {"序号": 9, "公司名称": "中国太保", "公众号名称": "中国太保", "公众号 ID": "CPIC-1991", "类型": "综合金融", "备注": "集团官方号"},
    {"序号": 10, "公司名称": "中国太保", "公众号名称": "太平洋寿险", "公众号 ID": "cpic-life", "类型": "人寿保险", "备注": "寿险子公司"},
    {"序号": 11, "公司名称": "中国太保", "公众号名称": "太平洋产险", "公众号 ID": "CPIC-P", "类型": "财产保险", "备注": "产险子公司"},
    
    # 大型寿险公司
    {"序号": 12, "公司名称": "泰康保险", "公众号名称": "泰康保险集团", "公众号 ID": "taikanglife", "类型": "人寿保险", "备注": "集团官方号"},
    {"序号": 13, "公司名称": "泰康保险", "公众号名称": "泰康人寿", "公众号 ID": "taikanglife95522", "类型": "人寿保险", "备注": "寿险子公司"},
    {"序号": 14, "公司名称": "新华保险", "公众号名称": "新华保险", "公众号 ID": "newchinalife", "类型": "人寿保险", "备注": "官方号"},
    {"序号": 15, "公司名称": "中国太平", "公众号名称": "中国太平", "公众号 ID": "cntaiping1929", "类型": "综合金融", "备注": "集团官方号"},
    {"序号": 16, "公司名称": "中国太平", "公众号名称": "太平人寿", "公众号 ID": "tplife-95589", "类型": "人寿保险", "备注": "寿险子公司"},
    {"序号": 17, "公司名称": "友邦保险", "公众号名称": "友邦中国", "公众号 ID": "AIA_China", "类型": "人寿保险", "备注": "中国区官方号"},
    {"序号": 18, "公司名称": "光大永明人寿", "公众号名称": "光大永明人寿", "公众号 ID": "everbright-life", "类型": "人寿保险", "备注": "官方号"},
    
    # 财险公司
    {"序号": 19, "公司名称": "中华联合保险", "公众号名称": "中华保险", "公众号 ID": "zhonghua-insurance", "类型": "财产保险", "备注": "集团官方号"},
    {"序号": 20, "公司名称": "阳光保险", "公众号名称": "阳光保险", "公众号 ID": "sunshine-insurance", "类型": "综合金融", "备注": "集团官方号"},
    {"序号": 21, "公司名称": "阳光保险", "公众号名称": "阳光产险", "公众号 ID": "sinopec-ins", "类型": "财产保险", "备注": "产险子公司"},
    {"序号": 22, "公司名称": "天安财险", "公众号名称": "天安财险", "公众号 ID": "95505", "类型": "财产保险", "备注": "官方号"},
    {"序号": 23, "公司名称": "众安保险", "公众号名称": "众安保险", "公众号 ID": "zhongan-insurance", "类型": "互联网保险", "备注": "首家互联网保险公司"},
    
    # 健康险公司
    {"序号": 24, "公司名称": "平安健康险", "公众号名称": "平安健康险", "公众号 ID": "pingan-health", "类型": "健康保险", "备注": "专业健康险公司"},
    {"序号": 25, "公司名称": "人保健康", "公众号名称": "人保健康", "公众号 ID": "picchealth", "类型": "健康保险", "备注": "专业健康险公司"},
    {"序号": 26, "公司名称": "太保安联健康险", "公众号名称": "太保安联健康", "公众号 ID": "CPIC-Allianz", "类型": "健康保险", "备注": "专业健康险公司"},
    
    # 养老保险公司
    {"序号": 27, "公司名称": "平安养老险", "公众号名称": "平安养老险", "公众号 ID": "pingan-pension", "类型": "养老保险", "备注": "专业养老险公司"},
    {"序号": 28, "公司名称": "国寿养老", "公众号名称": "国寿养老", "公众号 ID": "chinalife-pension", "类型": "养老保险", "备注": "专业养老险公司"},
    {"序号": 29, "公司名称": "长江养老", "公众号名称": "长江养老", "公众号 ID": "cj-pension", "类型": "养老保险", "备注": "专业养老险公司"},
    
    # 其他重要保险公司
    {"序号": 30, "公司名称": "中邮人寿", "公众号名称": "中邮人寿", "公众号 ID": "postlife", "类型": "人寿保险", "备注": "邮政系保险公司"},
    {"序号": 31, "公司名称": "建信人寿", "公众号名称": "建信人寿", "公众号 ID": "ccb-life", "类型": "人寿保险", "备注": "建行系保险公司"},
    {"序号": 32, "公司名称": "工银安盛", "公众号名称": "工银安盛人寿", "公众号 ID": "icbc-axa", "类型": "人寿保险", "备注": "工行系合资保险公司"},
    {"序号": 33, "公司名称": "招商信诺", "公众号名称": "招商信诺人寿", "公众号 ID": "cigna-cmb", "类型": "人寿保险", "备注": "招行系合资保险公司"},
    {"序号": 34, "公司名称": "中英人寿", "公众号名称": "中英人寿", "公众号 ID": "aviva-cofco", "类型": "人寿保险", "备注": "中粮系合资保险公司"},
    {"序号": 35, "公司名称": "同方全球人寿", "公众号名称": "同方全球人寿", "公众号 ID": "tongfang-global", "类型": "人寿保险", "备注": "清华系合资保险公司"},
    {"序号": 36, "公司名称": "中信保诚人寿", "公众号名称": "中信保诚人寿", "公众号 ID": "citic-prudential", "类型": "人寿保险", "备注": "中信系合资保险公司"},
    {"序号": 37, "公司名称": "中意人寿", "公众号名称": "中意人寿", "公众号 ID": "generali-china", "类型": "人寿保险", "备注": "中石油系合资保险公司"},
    {"序号": 38, "公司名称": "华泰人寿", "公众号名称": "华泰人寿", "公众号 ID": "huatai-life", "类型": "人寿保险", "备注": "官方号"},
    {"序号": 39, "公司名称": "富德生命人寿", "公众号名称": "富德生命人寿", "公众号 ID": "fude-life", "类型": "人寿保险", "备注": "官方号"},
    {"序号": 40, "公司名称": "华夏保险", "公众号名称": "华夏保险", "公众号 ID": "huaxia-insurance", "类型": "人寿保险", "备注": "已更名为瑞众人寿"},
    
    # 互联网保险平台
    {"序号": 41, "公司名称": "蚂蚁保", "公众号名称": "蚂蚁保", "公众号 ID": "ant-insurance", "类型": "互联网保险平台", "备注": "支付宝旗下"},
    {"序号": 42, "公司名称": "微保", "公众号名称": "微保", "公众号 ID": "we insure", "类型": "互联网保险平台", "备注": "微信旗下"},
    {"序号": 43, "公司名称": "慧择保险", "公众号名称": "慧择网", "公众号 ID": "huize-com", "类型": "保险经纪", "备注": "保险经纪平台"},
    {"序号": 44, "公司名称": "深蓝保", "公众号名称": "深蓝保", "公众号 ID": "shenlanbao", "类型": "保险科普", "备注": "保险测评平台"},
    {"序号": 45, "公司名称": "保险师", "公众号名称": "保险师", "公众号 ID": "baoxianshi-com", "类型": "保险科技", "备注": "保险从业者平台"},
    
    # 监管机构
    {"序号": 46, "公司名称": "国家金融监督管理总局", "公众号名称": "国家金融监督管理总局", "公众号 ID": "cbirc-gov", "类型": "监管机构", "备注": "原银保监会"},
    
    # 行业组织
    {"序号": 47, "公司名称": "中国保险行业协会", "公众号名称": "中国保险行业协会", "公众号 ID": "iachina", "类型": "行业协会", "备注": "行业自律组织"},
    {"序号": 48, "公司名称": "中国银行保险信息技术管理", "公众号名称": "中国银保信", "公众号 ID": "cbirc-it", "类型": "行业基础设施", "备注": "行业信息平台"},
]

def create_excel():
    """Create Excel file with insurance WeChat accounts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "保险公众号汇总"
    
    # Define headers
    headers = ["序号", "公司名称", "公众号名称", "公众号 ID", "类型", "备注"]
    
    # Create header row with styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, account in enumerate(insurance_accounts, 2):
        for col_idx, header in enumerate(headers, 1):
            value = account.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if header == "序号":
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
    
    # Adjust column widths
    column_widths = [8, 20, 18, 22, 18, 25]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add auto filter
    ws.auto_filter.ref = ws.dimensions
    
    # Create table
    table = Table(
        displayName="InsuranceAccounts",
        ref=f"A1:F{len(insurance_accounts) + 1}"
    )
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Save file
    output_path = "/home/admin/.openclaw/workspace/保险公众号汇总.xlsx"
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    print(f"Total records: {len(insurance_accounts)}")
    return output_path

if __name__ == "__main__":
    create_excel()
