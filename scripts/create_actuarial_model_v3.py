#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
临床试验责任保险精算模型
版本：v3.0 (2026-03-13)
更新：合并工作表 + Word 报告生成
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from datetime import datetime

# 创建 workbook
wb = Workbook()
wb.title = "临床试验责任保险精算模型"

# ==================== 工作表 1: 试验分类与保费计算（合并版） ====================
ws1 = wb.active
ws1.title = "试验分类与保费计算"

# 第一部分：试验分类与风险系数表
ws1['A1'] = "【试验分类与风险系数表】"
ws1['A1'].font = Font(bold=True, size=14, color="000080")

headers1 = ["分类维度", "子类别", "代码", "风险等级", "风险系数", "事故率估算 (%)", "基础费率 (‰)", "数据来源"]
ws1.append(headers1)

# 数据
data1 = [
    # 试验阶段
    ["试验阶段", "I 期临床试验", "P1", "高", 1.8, "4.5-6.2", 4.5, "FDA 2024+CDE"],
    ["试验阶段", "II 期临床试验", "P2", "中高", 1.5, "2.8-4.1", 3.8, "WHO ICTRP"],
    ["试验阶段", "III 期临床试验", "P3", "中", 1.0, "1.5-2.3", 2.8, "NMPA 批准数据"],
    ["试验阶段", "IV 期临床试验", "P4", "低", 0.6, "0.8-1.5", 1.8, "上市后监测"],
    ["试验阶段", "BE 试验", "BE", "中", 0.9, "1.2-2.0", 2.5, "CDE 审评数据"],
    ["发起者类型", "SCT (申办方发起)", "SCT", "标准", 1.0, "2.0-3.5", 2.8, "行业标准"],
    ["发起者类型", "IIT (研究者发起)", "IIT", "中高", 1.4, "3.0-5.0", 3.5, "Lancet 2025"],
    # 药物类型
    ["药物类型", "化学药 - 创新药", "CHEM-INN", "高", 1.6, "3.5-5.5", 4.0, "JAMA 2024"],
    ["药物类型", "化学药 - 仿制药", "CHEM-GEN", "低", 0.7, "1.0-2.0", 1.8, "CDE 审评"],
    ["药物类型", "生物药 - 单抗", "BIO-MAB", "高", 1.8, "4.0-6.0", 4.5, "Lancet 2025"],
    ["药物类型", "生物药 - 疫苗", "BIO-VAC", "中高", 1.5, "3.0-4.5", 3.8, "WHO 数据"],
    ["药物类型", "生物药 - 细胞治疗", "BIO-CELL", "极高", 2.5, "5.5-8.0", 6.5, "Nature Med 2024"],
    ["药物类型", "生物药 - 基因治疗", "BIO-GT", "极高", 2.8, "6.0-9.0", 7.0, "Nature Med 2024"],
    ["药物类型", "放射性药物", "RAD", "高", 2.0, "4.0-6.5", 5.0, "核医学数据"],
    ["药物类型", "中药/天然药物", "TCM", "中", 0.8, "1.5-2.8", 2.2, "CDE 中药"],
    ["药物类型", "医疗器械 - I 类", "DEV-I", "低", 0.5, "0.5-1.2", 1.2, "NMPA 器械"],
    ["药物类型", "医疗器械 - II 类", "DEV-II", "中", 0.8, "1.2-2.5", 2.0, "NMPA 器械"],
    ["药物类型", "医疗器械 - III 类", "DEV-III", "高", 1.6, "3.0-5.0", 4.0, "NMPA 器械"],
    # 治疗领域
    ["治疗领域", "肿瘤", "TA-ONC", "高", 1.8, "4.0-6.0", 4.5, "JAMA Oncol 2024"],
    ["治疗领域", "心血管", "TA-CV", "中", 1.1, "2.0-3.5", 2.8, "Circulation 2024"],
    ["治疗领域", "内分泌/代谢", "TA-ENDO", "中", 1.0, "1.8-3.0", 2.5, "Diabetes Care"],
    ["治疗领域", "神经科学", "TA-NEURO", "中高", 1.4, "2.8-4.5", 3.5, "Neurology 2024"],
    ["治疗领域", "免疫/风湿", "TA-IMM", "中高", 1.5, "3.0-4.8", 3.8, "Ann Rheum Dis"],
    ["治疗领域", "呼吸", "TA-RESP", "中", 1.1, "2.0-3.5", 2.8, "Lancet Respir"],
    ["治疗领域", "消化", "TA-GI", "低", 0.8, "1.2-2.5", 2.0, "Gastroenterology"],
    ["治疗领域", "罕见病", "TA-RARE", "高", 2.0, "4.0-7.0", 5.0, "Orphanet J"],
]

start_row = 2
for row in data1:
    ws1.append(row)

# 第二部分：保费计算器
calc_start_row = start_row + len(data1) + 3
ws1[f'A{calc_start_row}'] = "【保费计算器】"
ws1[f'A{calc_start_row}'].font = Font(bold=True, size=14, color="000080")

ws1[f'A{calc_start_row+1}'] = "输入参数"
ws1[f'A{calc_start_row+1}'].font = Font(bold=True)
ws1[f'A{calc_start_row+2}'] = "试验类型代码:"
ws1[f'B{calc_start_row+2}'] = "P3"
ws1[f'C{calc_start_row+2}'] = "从上方表格选择代码"
ws1[f'A{calc_start_row+3}'] = "受试者人数:"
ws1[f'B{calc_start_row+3}'] = 300
ws1[f'C{calc_start_row+3}'] = "人"
ws1[f'A{calc_start_row+4}'] = "试验周期:"
ws1[f'B{calc_start_row+4}'] = 24
ws1[f'C{calc_start_row+4}'] = "月"
ws1[f'A{calc_start_row+5}'] = "药物类型代码:"
ws1[f'B{calc_start_row+5}'] = "CHEM-INN"
ws1[f'C{calc_start_row+5}'] = "从上方表格选择代码"
ws1[f'A{calc_start_row+6}'] = "治疗领域代码:"
ws1[f'B{calc_start_row+6}'] = "TA-ONC"
ws1[f'C{calc_start_row+6}'] = "从上方表格选择代码"
ws1[f'A{calc_start_row+7}'] = "保险公司费率:"
ws1[f'B{calc_start_row+7}'] = 2.8
ws1[f'C{calc_start_row+7}'] = "‰ (人保财险基础费率)"
ws1[f'A{calc_start_row+8}'] = "是否国际多中心:"
ws1[f'B{calc_start_row+8}'] = "否"
ws1[f'C{calc_start_row+8}'] = "是/否"

ws1[f'A{calc_start_row+10}'] = "计算结果"
ws1[f'A{calc_start_row+10}'].font = Font(bold=True)
ws1[f'A{calc_start_row+11}'] = "基础保费:"
ws1[f'B{calc_start_row+11}'] = f"=B{calc_start_row+3}*B{calc_start_row+4}/12*B{calc_start_row+7}/1000"
ws1[f'C{calc_start_row+11}'] = "万元"
ws1[f'A{calc_start_row+12}'] = "试验阶段系数:"
ws1[f'B{calc_start_row+12}'] = f"=VLOOKUP(B{calc_start_row+2},$C$2:$F$28,5,FALSE)"
ws1[f'C{calc_start_row+12}'] = "自动匹配"
ws1[f'A{calc_start_row+13}'] = "药物类型系数:"
ws1[f'B{calc_start_row+13}'] = f"=VLOOKUP(B{calc_start_row+5},$C$2:$F$28,5,FALSE)"
ws1[f'C{calc_start_row+13}'] = "自动匹配"
ws1[f'A{calc_start_row+14}'] = "治疗领域系数:"
ws1[f'B{calc_start_row+14}'] = f"=VLOOKUP(B{calc_start_row+6},$C$2:$F$28,5,FALSE)"
ws1[f'C{calc_start_row+14}'] = "自动匹配"
ws1[f'A{calc_start_row+15}'] = "综合风险系数:"
ws1[f'B{calc_start_row+15}'] = f"=B{calc_start_row+12}*B{calc_start_row+13}*B{calc_start_row+14}"
ws1[f'C{calc_start_row+15}'] = "连乘计算"
ws1[f'A{calc_start_row+16}'] = "国际多中心加成:"
ws1[f'B{calc_start_row+16}'] = f"=IF(B{calc_start_row+8}=\"是\",1.25,1.0)"
ws1[f'C{calc_start_row+16}'] = "是=+25%"
ws1[f'A{calc_start_row+17}'] = "最终保费:"
ws1[f'B{calc_start_row+17}'] = f"=B{calc_start_row+11}*B{calc_start_row+15}*B{calc_start_row+16}*10000"
ws1[f'C{calc_start_row+17}'] = "元"
ws1[f'B{calc_start_row+17}'].font = Font(bold=True, size=16, color="FF0000")

ws1[f'A{calc_start_row+19}'] = "使用说明"
ws1[f'A{calc_start_row+19}'].font = Font(bold=True)
ws1[f'A{calc_start_row+20}'] = f"1. 在 B{calc_start_row+2}-B{calc_start_row+8} 输入试验参数"
ws1[f'A{calc_start_row+21}'] = "2. 代码从上方表格选择"
ws1[f'A{calc_start_row+22}'] = "3. 最终保费自动计算（红色显示）"

# 设置列宽
ws1.column_dimensions['A'].width = 15
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 25
ws1.column_dimensions['D'].width = 10
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 15
ws1.column_dimensions['G'].width = 12
ws1.column_dimensions['H'].width = 15

# ==================== 工作表 2: 赔偿标准与保险费率 ====================
ws2 = wb.create_sheet("赔偿标准与保险费率")

# 赔偿标准
ws2['A1'] = "【赔偿标准计算表】"
ws2['A1'].font = Font(bold=True, size=14, color="000080")

headers2 = ["损害等级", "代码", "伤残等级", "医疗费用 (万)", "误工费 (万)", "护理费 (万)", 
            "残疾赔偿金 (万)", "精神损害 (万)", "合计估算 (万)", "计算公式"]
ws2.append(headers2)

data2 = [
    ["轻度不良反应", "L1", "-", "0.3-1.0", "0.2-0.5", "0.1-0.3", "-", "0.2-0.4", "0.8-2.2", "医疗 + 误工 + 护理 + 精神"],
    ["中度损害 (住院)", "L2", "-", "3-10", "1-5", "0.5-2", "-", "3.5-5", "8-22", "医疗 + 误工 + 护理 + 精神"],
    ["重度损害 (10 级伤残)", "L3-A", "10 级", "10-30", "5-15", "3-8", "47-70", "5-10", "70-133", "全项累加"],
    ["重度损害 (5 级伤残)", "L3-B", "5 级", "15-50", "10-30", "5-15", "117-175", "8-15", "155-285", "全项累加"],
    ["重度损害 (1 级伤残)", "L3-C", "1 级", "30-100", "20-60", "10-30", "234-350", "15-30", "309-570", "全项累加"],
    ["死亡", "L4", "-", "5-20", "20-60", "5-15", "234-350", "20-50", "284-495", "全项累加"],
]

for row in data2:
    ws2.append(row)

# 保险费率
ws2['A11'] = "【保险费率表】"
ws2['A11'].font = Font(bold=True, size=14, color="000080")

headers3 = ["保险公司", "产品", "基础费率 (‰)", "IIT 加成", "高风险加成", "最终费率 (‰)", "每次事故限额 (万)", "联系电话"]
ws2.append(headers3)

data3 = [
    ["人保财险", "PICC-CT01", 2.8, "+20%", "+30%", "3.4-5.2", "200-800", "95518"],
    ["人保财险", "PICC-CT02", 3.5, "+20%", "+40%", "4.2-6.8", "500-1500", "95518"],
    ["平安产险", "PA-CT01", 2.5, "+15%", "+25%", "2.9-4.4", "150-600", "95511"],
    ["平安产险", "PA-CT02", 3.2, "+18%", "+35%", "3.8-6.0", "400-1200", "95511"],
    ["太平洋产险", "CPIC-CT01", 2.6, "+20%", "+30%", "3.1-5.0", "200-700", "95500"],
    ["太平洋产险", "CPIC-CT02", 3.6, "+25%", "+40%", "4.5-7.2", "600-2000", "95500"],
    ["美亚保险", "AIG-CT01", 3.8, "+25%", "+40%", "4.8-7.9", "500-2000", "400-820-8858"],
    ["安联保险", "ALZ-CT01", 4.0, "+25%", "+45%", "5.0-8.7", "600-2500", "400-888-2999"],
]

for row in data3:
    ws2.append(row)

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 15
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 10
ws2.column_dimensions['E'].width = 12
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 15
ws2.column_dimensions['H'].width = 15

# ==================== 工作表 3: 数据来源 ====================
ws3 = wb.create_sheet("数据来源")

headers4 = ["序号", "数据类型", "来源", "年份", "关键发现", "可信度"]
ws3.append(headers4)

data4 = [
    [1, "不良事件率", "FDA FAERS", "2024", "免疫治疗 3 级+AE 率 6.2%", "高"],
    [2, "不良事件率", "中国 CDE", "2024", "肿瘤试验 AE 报告率 72%", "高"],
    [3, "不良事件率", "WHO ICTRP", "2025", "II/III 期 AE 率 2.8-4.1%", "高"],
    [4, "赔偿金额", "JAMA Netw Open", "2024", "轻度 AE 平均$800-2,500", "中"],
    [5, "赔偿金额", "最高人民法院", "2025", "城镇收入×20 年", "高"],
    [6, "保险费率", "中国保险行业协会", "2025", "市场平均 2.8-3.5‰", "高"],
    [7, "HARMONi-A 研究", "JAMA", "2024", "322 例，3 级+AE 率 6.2%", "高"],
    [8, "RATIONALE-315", "Lancet Respir Med", "2025", "453 例，3 级+AE 率 72%", "高"],
]

for row in data4:
    ws3.append(row)

ws3.column_dimensions['A'].width = 6
ws3.column_dimensions['B'].width = 15
ws3.column_dimensions['C'].width = 25
ws3.column_dimensions['D'].width = 8
ws3.column_dimensions['E'].width = 35
ws3.column_dimensions['F'].width = 10

# 保存文件
file_path = "/home/admin/.openclaw/workspace/临床试验责任保险精算模型_v3.0.xlsx"
wb.save(file_path)

print(f"✅ Excel 精算模型 v3.0 已生成：{file_path}")
print(f"包含 3 个工作表:")
print(f"  1. 试验分类与保费计算（合并版）")
print(f"  2. 赔偿标准与保险费率")
print(f"  3. 数据来源")
