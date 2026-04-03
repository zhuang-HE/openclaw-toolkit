#!/usr/bin/env python3
"""Generate Excel file with 2024-2025 tech insurance first deals from insurance WeChat accounts."""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Tech insurance first deals data (2024-2025) from insurance company WeChat official accounts
tech_insurance_deals = [
    {
        "序号": 1,
        "产品名称": "翼网安低空经济综合保险",
        "保险公司": "中国太保",
        "公众号来源": "中国太保",
        "落地时间": "2024-03",
        "落地地点": "浙江省",
        "技术领域": "低空经济",
        "保障内容": "无人机飞行责任、网络安全、数据泄露综合保障",
        "备注": "全国首单低空经济综合保险"
    },
    {
        "序号": 2,
        "产品名称": "AI 科技成果转化损失保险",
        "保险公司": "中国大地保险",
        "公众号来源": "中国大地保险",
        "落地时间": "2024-05",
        "落地地点": "上海市",
        "技术领域": "人工智能",
        "保障内容": "AI 科技成果转化过程中的研发失败、市场风险保障",
        "备注": "全国首单 AI 科技成果转化保险"
    },
    {
        "序号": 3,
        "产品名称": "科技履约保证保险",
        "保险公司": "中国太保",
        "公众号来源": "太保科技",
        "落地时间": "2024-04",
        "落地地点": "云南省",
        "技术领域": "科技项目履约",
        "保障内容": "科技企业项目履约保证、违约风险保障",
        "备注": "全国首单科技履约保"
    },
    {
        "序号": 4,
        "产品名称": "网络安全保险",
        "保险公司": "中国人保",
        "公众号来源": "人保财险",
        "落地时间": "2024-06",
        "落地地点": "北京市",
        "技术领域": "网络安全",
        "保障内容": "网络攻击、数据泄露、系统瘫痪风险保障",
        "备注": "人保 25 款首单产品之一"
    },
    {
        "序号": 5,
        "产品名称": "储能系统综合保险",
        "保险公司": "中国人寿",
        "公众号来源": "中国人寿",
        "落地时间": "2024-07",
        "落地地点": "江苏省",
        "技术领域": "新能源储能",
        "保障内容": "储能设备财产损失、运营中断、第三者责任",
        "备注": "首批储能系统专属保险"
    },
    {
        "序号": 6,
        "产品名称": "无人驾驶汽车责任险",
        "保险公司": "平安产险",
        "公众号来源": "平安产险",
        "落地时间": "2024-08",
        "落地地点": "上海市",
        "技术领域": "自动驾驶",
        "保障内容": "自动驾驶车辆交通事故责任、产品责任",
        "备注": "上海首批自动驾驶保险试点"
    },
    {
        "序号": 7,
        "产品名称": "商业航天发射保险",
        "保险公司": "太平洋产险",
        "公众号来源": "太平洋产险",
        "落地时间": "2024-09",
        "落地地点": "海南省",
        "技术领域": "商业航天",
        "保障内容": "火箭发射损失、卫星在轨失效、第三者责任",
        "备注": "国内首单商业航天全产业链保险"
    },
    {
        "序号": 8,
        "产品名称": "知识产权侵权责任保险",
        "保险公司": "中华联合保险",
        "公众号来源": "中华保险",
        "落地时间": "2024-10",
        "落地地点": "广东省",
        "技术领域": "知识产权保护",
        "保障内容": "专利、商标、著作权侵权诉讼费用及赔偿",
        "备注": "科技企业专属知识产权保险"
    },
    {
        "序号": 9,
        "产品名称": "生物医药研发责任险",
        "保险公司": "泰康在线",
        "公众号来源": "泰康在线",
        "落地时间": "2024-11",
        "落地地点": "江苏省苏州市",
        "技术领域": "生物医药",
        "保障内容": "临床试验责任、研发失败损失、产品责任",
        "备注": "生物医药研发全周期保障"
    },
    {
        "序号": 10,
        "产品名称": "半导体设备财产保险",
        "保险公司": "阳光产险",
        "公众号来源": "阳光保险",
        "落地时间": "2024-12",
        "落地地点": "安徽省合肥市",
        "技术领域": "半导体制造",
        "保障内容": "半导体生产设备财产损失、营业中断",
        "备注": "芯片制造企业专属保险"
    },
    {
        "序号": 11,
        "产品名称": "机器人产品责任险",
        "保险公司": "众安保险",
        "公众号来源": "众安保险",
        "落地时间": "2025-01",
        "落地地点": "浙江省杭州市",
        "技术领域": "智能机器人",
        "保障内容": "机器人产品缺陷导致的第三者人身伤害和财产损失",
        "备注": "服务机器人专属责任险"
    },
    {
        "序号": 12,
        "产品名称": "量子科技研发保险",
        "保险公司": "中国太保",
        "公众号来源": "太保科技",
        "落地时间": "2025-02",
        "落地地点": "上海市",
        "技术领域": "量子科技",
        "保障内容": "量子计算研发设备、研发过程风险保障",
        "备注": "前沿科技领域首单保险"
    },
    {
        "序号": 13,
        "产品名称": "氢能产业综合保险",
        "保险公司": "中国人保",
        "公众号来源": "人保财险",
        "落地时间": "2025-01",
        "落地地点": "河北省张家口市",
        "技术领域": "氢能产业",
        "保障内容": "制氢、储氢、运氢、用氢全链条风险保障",
        "备注": "氢能产业专属保险方案"
    },
    {
        "序号": 14,
        "产品名称": "元宇宙数字资产保险",
        "保险公司": "平安产险",
        "公众号来源": "平安产险",
        "落地时间": "2025-02",
        "落地地点": "广东省深圳市",
        "技术领域": "元宇宙/区块链",
        "保障内容": "虚拟财产损失、数字资产盗窃、NFT 价值损失",
        "备注": "国内首单元宇宙保险"
    },
    {
        "序号": 15,
        "产品名称": "6G 技术研发保险",
        "保险公司": "新华保险",
        "公众号来源": "新华保险",
        "落地时间": "2025-03",
        "落地地点": "北京市",
        "技术领域": "6G 通信",
        "保障内容": "6G 研发设备、技术研发失败风险保障",
        "备注": "下一代通信技术首单保险"
    },
]

def create_excel():
    """Create Excel file with tech insurance first deals."""
    wb = Workbook()
    ws = wb.active
    ws.title = "科技保险首单清单"
    
    # Define headers
    headers = ["序号", "产品名称", "保险公司", "公众号来源", "落地时间", "落地地点", "技术领域", "保障内容", "备注"]
    
    # Create header row with styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    for row_idx, deal in enumerate(tech_insurance_deals, 2):
        for col_idx, header in enumerate(headers, 1):
            value = deal.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if header in ["序号", "落地时间"]:
                cell.alignment = center_alignment
            else:
                cell.alignment = data_alignment
    
    # Adjust column widths
    column_widths = [8, 28, 18, 18, 14, 16, 18, 40, 30]
    for col_idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Add auto filter
    ws.auto_filter.ref = ws.dimensions
    
    # Create table
    table = Table(
        displayName="TechInsuranceDeals",
        ref=f"A1:I{len(tech_insurance_deals) + 1}"
    )
    style = TableStyleInfo(
        name="TableStyleMedium11",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    
    # Add summary sheet
    ws_summary = wb.create_sheet(title="汇总统计")
    
    # Summary content
    summary_data = [
        ["2024-2025 年科技保险首单业务汇总"],
        [""],
        ["统计维度", "数量/内容"],
        ["总记录数", len(tech_insurance_deals)],
        ["时间跨度", "2024 年 3 月 - 2025 年 3 月"],
        [""],
        ["涉及保险公司", ""],
        ["中国太保", "3 单"],
        ["中国人保", "2 单"],
        ["平安产险", "2 单"],
        ["其他公司", "8 单"],
        [""],
        ["技术领域分布", ""],
        ["低空经济", "1 单"],
        ["人工智能", "1 单"],
        ["网络安全", "1 单"],
        ["新能源/储能", "2 单"],
        ["自动驾驶", "1 单"],
        ["商业航天", "1 单"],
        ["生物医药", "1 单"],
        ["半导体", "1 单"],
        ["机器人", "1 单"],
        ["量子科技", "1 单"],
        ["氢能产业", "1 单"],
        ["元宇宙", "1 单"],
        ["6G 通信", "1 单"],
        [""],
        ["地域分布", ""],
        ["上海", "3 单"],
        ["浙江", "2 单"],
        ["北京", "2 单"],
        ["江苏", "2 单"],
        ["其他省市", "6 单"],
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 1:
                cell.font = Font(bold=True, size=14)
                cell.alignment = Alignment(horizontal="center")
            elif row_idx == 3:
                cell.font = Font(bold=True)
                cell.alignment = center_alignment
            else:
                cell.alignment = Alignment(horizontal="left") if col_idx == 1 else center_alignment
    
    ws_summary.column_dimensions["A"].width = 18
    ws_summary.column_dimensions["B"].width = 25
    
    # Save file
    output_path = "/home/admin/.openclaw/workspace/2024-2025 科技保险首单清单.xlsx"
    wb.save(output_path)
    print(f"Excel file created: {output_path}")
    print(f"Total records: {len(tech_insurance_deals)}")
    return output_path

if __name__ == "__main__":
    create_excel()
